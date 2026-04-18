from __future__ import annotations

from pathlib import Path

from fastapi import HTTPException


def resolve_local_path_from_files_url(file_url: str) -> Path:
    """Only allow /files/<name> served by this backend to avoid SSRF/path traversal."""
    if not isinstance(file_url, str) or not file_url.strip():
        raise HTTPException(status_code=400, detail="fileUrl required")

    url = file_url.strip()
    if not url.startswith("/files/"):
        raise HTTPException(status_code=400, detail="fileUrl must start with /files/")

    rel = url[len("/files/") :]
    if not rel:
        raise HTTPException(status_code=400, detail="invalid fileUrl")

    # Allow a single subdirectory (e.g. voice/<name>) but prevent traversal.
    parts = [p for p in rel.split("/") if p]
    if len(parts) > 2:
        raise HTTPException(status_code=400, detail="invalid fileUrl")
    if any(p in ("..", ".") for p in parts):
        raise HTTPException(status_code=400, detail="invalid fileUrl")

    base = Path("/tmp/uploads")
    fp = base.joinpath(*parts)
    if not fp.exists() or not fp.is_file():
        raise HTTPException(status_code=404, detail="file not found")

    return fp


def extract_text(fp: Path, mime: str | None = None) -> tuple[str, dict]:
    """Extract plain text from common document formats.

    Returns: (text, meta)
    - meta may include pages, type, etc.

    Raises HTTPException for unsupported/parse errors.
    """

    ext = fp.suffix.lower()
    # allow mime hint but don't trust it fully

    if ext == ".pdf" or (mime == "application/pdf"):
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(fp))
            pages = reader.pages
            texts = []
            for p in pages:
                try:
                    texts.append(p.extract_text() or "")
                except Exception:
                    texts.append("")
            text = "\n".join([t for t in texts if t])
            return text, {"pages": len(pages), "type": "pdf"}
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=422, detail="pdf parse failed")

    if ext in (".docx",) or (
        mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        try:
            import docx  # python-docx

            d = docx.Document(str(fp))
            parts = [p.text for p in d.paragraphs if p.text]
            text = "\n".join(parts)
            return text, {"type": "docx"}
        except Exception:
            raise HTTPException(status_code=422, detail="docx parse failed")

    if ext in (".xlsx", ".xls") or (
        mime
        in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        )
    ):
        # Prefer openpyxl for xlsx. For xls (legacy), we fall back to a friendly error unless xlrd is added.
        if ext == ".xls" and mime == "application/vnd.ms-excel":
            raise HTTPException(status_code=415, detail="xls is not supported yet; please upload xlsx")
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.worksheets:
                lines.append(f"# sheet: {sheet.title}")
                # limit rows/cols to avoid giant prompts
                max_rows = min(sheet.max_row or 0, 200)
                max_cols = min(sheet.max_column or 0, 30)
                for r in range(1, max_rows + 1):
                    row_vals = []
                    for c in range(1, max_cols + 1):
                        v = sheet.cell(row=r, column=c).value
                        if v is None:
                            row_vals.append("")
                        else:
                            s = str(v).strip()
                            # avoid very long cell
                            if len(s) > 200:
                                s = s[:200] + "..."
                            row_vals.append(s)
                    # skip empty rows
                    if any(x for x in row_vals):
                        lines.append("\t".join(row_vals))
            text = "\n".join(lines).strip()
            return text, {"type": "xlsx", "sheets": len(wb.worksheets)}
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=422, detail="xlsx parse failed")

    if ext in (".txt", ".md") or (mime and mime.startswith("text/")):
        try:
            text = fp.read_text(encoding="utf-8", errors="ignore")
            return text, {"type": "text"}
        except Exception:
            raise HTTPException(status_code=422, detail="text parse failed")

    raise HTTPException(status_code=415, detail="unsupported document type")
